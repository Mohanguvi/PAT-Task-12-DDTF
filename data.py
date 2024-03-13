from openpyxl.reader.excel import load_workbook


class WebPage:
    """
        To perform the testing for the OrangeHRM
        The test data are given from the DDF.xlsx Excel file.
       """

    def __init__(self):
        """
            The __init__() method initializes the class instance with the following attributes:
            - url: The URL of the web page.
            - LoginUrl: The URL of the login page.
            - excelFile: The path to an Excel file containing data.
            - sheetName: The name of the worksheet within the Excel file.
            - book: Loads the Excel file using the openpyxl library and assigns it to the book attribute.
            - file: Accesses the specific worksheet within the Excel file and assigns it to the file attribute.
        """
        self.url = "https://opensource-demo.orangehrmlive.com/web/index.php/auth/login"
        self.LoginUrl = "https://opensource-demo.orangehrmlive.com/web/index.php/dashboard/index"
        self.excelFile = "Data/DDF.xlsx"
        self.sheetName = "Sheet1"
        self.book = load_workbook(self.excelFile)
        self.file = self.book[self.sheetName]

    def rowTotal(self):
        """
        The rowTotal() method returns the total number of rows in the worksheet.
        :return: max data count of row
        """
        return self.file.max_row

    def readExcel(self, row, column):
        """
        The readExcel() method takes a row and column as arguments and
         returns the value of the cell at that position in the worksheet.
        :param row:
        :param column:
        :return:
        """
        return self.file.cell(row, column).value

    def writeExcel(self, row, column, data):
        """
        The writeExcel() method takes a row, column, and data as arguments.
        It updates the value of the cell at the specified row and column with the provided data and
        saves the changes to the Excel file.
        :param row:
        :param column:
        :param data:
        :return:
        """
        self.file.cell(row, column).value = data
        self.book.save(self.excelFile)

